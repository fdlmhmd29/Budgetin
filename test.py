import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
import locale
import os

# Set locale Indonesia
try:
    locale.setlocale(locale.LC_TIME, 'id_ID.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_TIME, 'Indonesian')
    except:
        pass

# Tarif harian
tarif = {
    'senin': {'krl': 8000, 'krd': 4000},
    'selasa': {'krl': 6000, 'krd': 0},
    'rabu': {'krl': 6000, 'krd': 0},
    'kamis': {'krl': 6000, 'krd': 0},
    'jumat': {'krl': 8000, 'krd': 3000},
    'sabtu': {'krl': 0, 'krd': 0},
    'minggu': {'krl': 0, 'krd': 0}
}

map_hari = {
    0: 'senin',
    1: 'selasa',
    2: 'rabu',
    3: 'kamis',
    4: 'jumat',
    5: 'sabtu',
    6: 'minggu'
}

selected_folder = ""  # Folder yang dipilih

def pilih_folder():
    global selected_folder
    folder = filedialog.askdirectory()
    if folder:
        selected_folder = folder
        folder_label.config(text=f"Folder: {folder}")

def generate_budget_gui(bulan_awal, tahun):
    try:
        if not selected_folder:
            messagebox.showwarning("Folder Belum Dipilih", "Silakan pilih folder penyimpanan terlebih dahulu.")
            return

        bulan_awal = int(bulan_awal)
        tahun = int(tahun)
        start_date = datetime(tahun, bulan_awal, 29)

        if bulan_awal == 12:
            bulan_akhir = 1
            tahun_akhir = tahun + 1
        else:
            bulan_akhir = bulan_awal + 1
            tahun_akhir = tahun

        end_date = datetime(tahun_akhir, bulan_akhir, 28)

        total_krl = 0
        total_krd = 0
        jumlah_hari = {h: 0 for h in tarif}
        log_harian = []
        mingguan = []
        week_num = 1
        week_start = start_date

        current = start_date
        while current <= end_date:
            kode_hari = current.weekday()
            hari = map_hari[kode_hari]
            krl = tarif[hari]['krl']
            krd = tarif[hari]['krd']

            total_krl += krl
            total_krd += krd
            jumlah_hari[hari] += 1

            log_harian.append({
                'Tanggal': current.strftime('%Y-%m-%d'),
                'Hari': hari.capitalize(),
                'Biaya KRL (Rp)': krl,
                'Biaya KRD (Rp)': krd,
                'Total Harian (Rp)': krl + krd
            })

            # Rekap mingguan tiap Sabtu atau akhir
            if current.weekday() == 5 or current == end_date:
                mingguan.append({
                    'Minggu ke-': week_num,
                    'Tanggal': f"{week_start.strftime('%d %b')} - {current.strftime('%d %b')}",
                    'Total KRL': sum(row['Biaya KRL (Rp)'] for row in log_harian if week_start.strftime('%Y-%m-%d') <= row['Tanggal'] <= current.strftime('%Y-%m-%d')),
                    'Total KRD': sum(row['Biaya KRD (Rp)'] for row in log_harian if week_start.strftime('%Y-%m-%d') <= row['Tanggal'] <= current.strftime('%Y-%m-%d')),
                })
                week_num += 1
                week_start = current + timedelta(days=1)

            current += timedelta(days=1)

        total = total_krl + total_krd

        nama_file = f"Budgeting_Transport_{start_date.strftime('%d %B')}_{end_date.strftime('%d %B')}.xlsx"
        full_path = os.path.join(selected_folder, nama_file)
        sheet_name = f"{start_date.strftime('%b')} - {end_date.strftime('%b')}"

        # Tulis ke Excel
        df = pd.DataFrame(log_harian)
        df_week = pd.DataFrame(mingguan)
        with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            df_week.to_excel(writer, index=False, sheet_name="Rekap Mingguan")

        # Format Excel
        wb = openpyxl.load_workbook(full_path)
        ws = wb[sheet_name]

        header_fill = PatternFill("solid", fgColor="B7DEE8")
        friday_fill = PatternFill("solid", fgColor="FFD966")
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
            if row[1].value == "Jumat":
                for cell in row:
                    cell.fill = friday_fill
            row[2].number_format = '#,##0'
            row[3].number_format = '#,##0'
            row[4].number_format = '#,##0'

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        ws.freeze_panes = 'A2'

        # Total bawah
        total_row = ws.max_row + 1
        ws[f"B{total_row}"] = "TOTAL"
        ws[f"B{total_row}"].font = bold_font
        for col in range(3, 6):
            col_letter = get_column_letter(col)
            ws[f"{col_letter}{total_row}"] = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
            ws[f"{col_letter}{total_row}"].font = bold_font
            ws[f"{col_letter}{total_row}"].number_format = '#,##0'

        # Chart batang
        chart = BarChart()
        chart.title = "Biaya Transport Harian"
        chart.y_axis.title = "Biaya (Rp)"
        chart.x_axis.title = "Tanggal"
        data = Reference(ws, min_col=3, max_col=5, min_row=1, max_row=ws.max_row - 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "H2")

        # Pie chart
        ws_summary = wb.create_sheet("Rekap Total")
        ws_summary.append(["Kategori", "Total"])
        ws_summary.append(["KRL", total_krl])
        ws_summary.append(["KRD", total_krd])
        pie = PieChart()
        pie.add_data(Reference(ws_summary, min_col=2, min_row=2, max_row=3))
        pie.set_categories(Reference(ws_summary, min_col=1, min_row=2, max_row=3))
        pie.title = "Distribusi Biaya KRL vs KRD"
        ws_summary.add_chart(pie, "D2")

        wb.save(full_path)

        hasil = f"""Periode: {start_date.strftime('%d %B %Y')} - {end_date.strftime('%d %B %Y')}
Total KRL: Rp{total_krl:,}
Total KRD: Rp{total_krd:,}
Total: Rp{total:,}
File: {full_path}
"""
        messagebox.showinfo("Hasil Budgeting", hasil)

    except Exception as e:
        messagebox.showerror("Error", f"Terjadi kesalahan: {e}")

# GUI setup
root = tk.Tk()
root.title("Budgeting Transport GUI")
root.geometry("480x300")

tk.Label(root, text="Pilih Bulan Awal:").pack(pady=5)
bulan_var = ttk.Combobox(root, values=[str(i) for i in range(1, 13)], state="readonly")
bulan_var.current(datetime.now().month - 1)
bulan_var.pack()

tk.Label(root, text="Masukkan Tahun:").pack(pady=5)
tahun_entry = tk.Entry(root)
tahun_entry.insert(0, str(datetime.now().year))
tahun_entry.pack()

tk.Button(root, text="Pilih Folder Penyimpanan", command=pilih_folder).pack(pady=10)
folder_label = tk.Label(root, text="Folder: (belum dipilih)")
folder_label.pack()

def on_generate():
    bulan = bulan_var.get()
    tahun = tahun_entry.get()
    if not bulan or not tahun:
        messagebox.showwarning("Input Salah", "Silakan isi bulan dan tahun.")
        return
    generate_budget_gui(bulan, tahun)

tk.Button(root, text="Generate Budget", command=on_generate, bg="#4CAF50", fg="white").pack(pady=20)

root.mainloop()
